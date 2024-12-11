
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
import random
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os  
from PyPDF2 import PdfMerger  # Import PdfMerger from PyPDF2
# excel file generate
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from flask import send_file
import os
import io


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timetable.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key_here'  # Required for flash messages
db = SQLAlchemy(app)

# Database Models
class Classroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)

class Lab(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)

class Batch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    credits = db.Column(db.Integer, nullable=False)
    is_lab = db.Column(db.Boolean, default=False)
    priority = db.Column(db.Boolean, default=False)

class Professor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)

class Schedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=False)
    day = db.Column(db.Integer, nullable=False)  # 0-4 for Monday-Friday
    slot = db.Column(db.Integer, nullable=False)  # 0-8 for time slots
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=True)
    lab_id= db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)

# roomallotment 
def find_available_classroom(day, slot):
    all_classrooms = Classroom.query.all()
    for classroom in all_classrooms:
        existing_schedule = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot
        ).first()
        if not existing_schedule:
            return classroom
    return None

def find_available_lab(day, slot):
    all_classrooms = Lab.query.all()
    for lab in all_classrooms:
        # Check the current slot
        current_schedule = Schedule.query.filter_by(
            classroom_id=lab.id, day=day, slot=slot
        ).first()
        
        # Check the next slot
        next_slot = Schedule.query.filter_by(
            classroom_id=lab.id, day=day, slot=slot + 1
        ).first()
        
        # Check the slot after next
        next_next_slot = Schedule.query.filter_by(
            classroom_id=lab.id, day=day, slot=slot + 2
        ).first()
        
        # If all three slots are free, return the lab
        if not current_schedule and not next_slot and not next_next_slot:
            return lab

    # If no lab is available for all three slots
    return None


# Helper function to check if a slot is available
def is_slot_available(batch_id, professor_id, day, slot, is_lab=False):
    if is_lab:
        # Check for three consecutive slots for a lab
        if slot > 6 or slot == 2 or slot == 4 or slot ==3 :  # Can't start a lab if the starting slot is beyond 6 (not enough space for 3 slots)
            return False
        for offset in range(3):
            existing_schedule = Schedule.query.filter_by(
                batch_id=batch_id, day=day, slot=slot + offset
            ).first()
            professor_schedule = Schedule.query.filter_by(
                professor_id=professor_id, day=day, slot=slot + offset
            ).first()
            if existing_schedule or professor_schedule:
                return False
        return True
    else:
        # Check single slot availability for non-lab courses
        existing_schedule = Schedule.query.filter_by(
            batch_id=batch_id, day=day, slot=slot
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=professor_id, day=day, slot=slot
        ).first()
        return not (existing_schedule or professor_schedule)






# excel code

def download_excel_files():
    # Create a new workbook to store the merged result
    batches = Batch.query.all()  # Retrieve all batches from the database
    
    # Loop through each batch and generate Excel
    for batch in batches:
        batch_id = batch.id
        generate_excel(batch_id)

def generate_excel(batch_ids):
    # Fetch the batches based on provided IDs
    selected_batches = Batch.query.filter(Batch.id.in_(batch_ids)).all()
    if not selected_batches:
        return "No valid batches found for the provided IDs."

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM", "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    excel_paths = []  # To store the paths of generated Excel files
    timetable = [[" " for _ in range(10)] for _ in range(5)]
    for batch in selected_batches:
        # Initialize the timetable with empty values
        
        schedules = Schedule.query.filter_by(batch_id=batch.id).all()

        # Fill in the timetable based on schedules
        for schedule in schedules:
            course = Course.query.get(schedule.course_id)
            professor = Professor.query.get(schedule.professor_id)
            classroom = Classroom.query.get(schedule.classroom_id)
            lab = Lab.query.get(schedule.lab_id)
            entry = f"{course.name} ({professor.name})"
            if lab:
                entry += f", {lab.name}"
            if classroom:
                entry += f" {classroom.name}"
            timetable[schedule.day][schedule.slot] += "\n" + entry

     # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"TimeTable"

        # Set the title row
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        
        # Set header row
        ws['A1'] = f"Timetable for Batch: {batch.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:K1')  # Merge columns to fit time slots

        # Adding day and time slot headers
        ws['A2'] = "Day"
        for col_index, time_slot in enumerate(time_slots, start=2):
            col_letter = get_column_letter(col_index)
            ws[f'{col_letter}2'] = time_slot

        # Filling in the timetable data
        for day_index, day_name in enumerate(days):
            ws[f"A{day_index + 3}"] = day_name
            for slot_index, slot_content in enumerate(timetable[day_index]):
                col_letter = get_column_letter(slot_index + 2)
                ws[f"{col_letter}{day_index + 3}"] = slot_content

        # Save the Excel file to a path
    # excel_path = f"C:\Users\ASUS\Downloads\Time-Table-project (8)\Time-Table-project/timetables/batch_{batch.id}.xlsx"
    excel_path = "C:\\Users\\ASUS\\Downloads\\Time-Table-project (8)\\Time-Table-project\\timetables\\batch_{batch.id}.xlsx"


    wb.save(excel_path)
        

    return excel_path

# excel file end

# Routes
@app.route('/')
def index():
    batches = Batch.query.all()
    return render_template('index.html', batches=batches)

@app.route('/select-batches', methods=['GET'])
def select_batches():
    # Fetch all batches from the database
    batches = Batch.query.all()
    return render_template('select_batches.html', batches=batches)

@app.route('/download-timetable', methods=['POST'])
def download_timetable():
    # Get the selected batch IDs from the form
    selected_batch_ids = request.form.getlist('batch_ids[]')
    if not selected_batch_ids:
        return "No batches selected. Please select at least one batch."

    # Generate timetables for the selected batches
    excel_path = generate_excel(selected_batch_ids)

    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True)
    else:
        return "File not found", 404


@app.route('/create_batch', methods=['GET', 'POST'])
def create_batch():
    if request.method == 'POST':
        batch_name = request.form['name']
        new_batch = Batch(name=batch_name)
        db.session.add(new_batch)
        try:
            db.session.commit()
            flash('Batch created successfully', 'success')
        except:
            db.session.rollback()
            flash('Error creating batch', 'error')
        return redirect(url_for('index'))
    return render_template('create_batch.html')
# comment here caution

# classroom route
@app.route('/classrooms', methods=['GET', 'POST'])
def manage_classrooms():
    if request.method == 'POST':
        classroom_name = request.form['name']
        capacity = int(request.form['capacity'])
        new_classroom = Classroom(name=classroom_name, capacity=capacity)
        db.session.add(new_classroom)
        try:
            db.session.commit()
            flash('Classroom added successfully', 'success')
        except:
            db.session.rollback()
            flash('Error adding classroom', 'error')
        return redirect(url_for('manage_classrooms'))
    classrooms = Classroom.query.all()
    return render_template('classrooms.html', classrooms=classrooms)

@app.route('/labs', methods=['GET', 'POST'])
def manage_labs():
    if request.method == 'POST':
        lab_name = request.form['lab_name']
        lab_capacity = int(request.form['lab_capacity'])
        new_lab = Lab(name=lab_name, capacity=lab_capacity)
        db.session.add(new_lab)
        try:
            db.session.commit()
            flash('Lab added successfully', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding lab: {e}', 'error')
        return redirect(url_for('manage_labs'))
    
    labs = Lab.query.all()
    return render_template('labs.html', labs=labs)

# classroom route end

@app.route('/batch/<int:batch_id>', methods=['GET', 'POST'])
def manage_batch(batch_id):
    batch = Batch.query.get_or_404(batch_id)
    if request.method == 'POST':
        course_name = request.form['course_name']
        credits = int(request.form['credits'])
        is_lab = 'is_lab' in request.form
        priority = 'priority' in request.form
        priority_type = request.form.get('priority_type')  # Get the priority type
        priority_shift='priority_shift' in request.form
        priority_shift_type=request.form.get('priority_shift_type')
        priority_day='priority_day' in request.form
        priority_day_type=request.form.get('priority_day_type')
        

        professor_name = request.form['professor_name']
        course = Course.query.filter_by(name=course_name).first()
        if not course:
            course = Course(name=course_name, credits=credits, is_lab=is_lab, priority=priority)
            db.session.add(course)

        professor = Professor.query.filter_by(name=professor_name).first()
        if not professor:
            professor = Professor(name=professor_name)
            db.session.add(professor)

        try:
            db.session.commit()
        except:
            db.session.rollback()
            flash('Error adding course or professor', 'error')
            return redirect(url_for('manage_batch', batch_id=batch_id))

        # Schedule the course based on priority type
        day_map = {
            "Monday": 0,
            "Tuesday": 1,
            "Wednesday": 2,
            "Thursday": 3,
            "Friday": 4
        }
        priority_day_index = day_map.get(priority_day_type, -1)
        
        available_slots = []
        if is_lab:
            for day in range(5):
                for slot in range(7):  # Only check up to slot 6 for labs
                    if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot!=4:
                        available_slots.append((day, slot))
                        break
        elif priority_day:
            for day in range(5):  
                if day == priority_day_type:
                    for slot in range(9):  
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))  
                            break  
                else:
                    for slot in range(5, 9):  
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))  
                            break 
        

                
        elif priority_shift:
            if priority_shift_type=="first_half":
                for day in range(5):
                    for slot in range(5):
                        if is_slot_available(batch.id,professor.id,day,slot,is_lab=True) and slot!=4:
                            available_slots.append((day,slot))
                            break
            else:
                for day in range(5):
                    for slot in range(5,9):
                        if is_slot_available(batch.id,professor.id,day,slot,is_lab=True) and slot!=4:
                            available_slots.append((day,slot))
                            break
        elif priority:
            if priority_type == "2-hour consecutive":
                for day in range(5):
                    for slot in range(8):  # Check up to slot 7 for consecutive slots
                        if slot != 3 and slot !=4:
                            if is_slot_available(batch.id, professor.id, day, slot) and is_slot_available(batch.id, professor.id, day, slot + 1):
                                available_slots.append((day, slot))
                                break
            elif priority_type == "2-1-1":
                for day in range(5):
                    for slot in range(8):  # Check up to slot 7 for consecutive slots
                        if slot != 3 and slot !=4:
                            if is_slot_available(batch.id, professor.id, day, slot) and is_slot_available(batch.id, professor.id, day, slot + 1) and slot!=4:
                                available_slots.append((day, slot))
                                break
                credits-=2
                for day in range(5):
                    for slot in range(9):  # Check all slots for 2-1-1 hours pattern
                        if is_slot_available(batch.id, professor.id, day, slot) and slot!=4 and credits>0:
                            available_slots.append((day, slot))
                            break
                credits-=1
                for day in range(5):
                    for slot in range(9):  # Check all slots for 2-1-1 hours pattern
                        if is_slot_available(batch.id, professor.id, day, slot) and slot!=4 and credits>0:
                            available_slots.append((day, slot))
                            break
            elif priority_type == "2-1":
                for day in range(5):
                    for slot in range(8):  # Check up to slot 7 for consecutive slots
                        if slot != 3 and slot !=4:
                            if is_slot_available(batch.id, professor.id, day, slot) and is_slot_available(batch.id, professor.id, day, slot + 1) and slot!=4:
                                available_slots.append((day, slot))
                                break
                credits-=2
                for day in range(5):
                    for slot in range(9):  # Check all slots for 2-1-1 hours pattern
                        if is_slot_available(batch.id, professor.id, day, slot) and slot!=4 and credits>0:
                            available_slots.append((day, slot))
                            break
        else:
            for day in range(5):
                for slot in range(9):
                    if is_slot_available(batch.id, professor.id, day, slot) and slot!=4:
                        available_slots.append((day, slot))
                        break

        if priority and priority_type == "2-hour consecutive":
            if not available_slots:
                flash('Not enough consecutive slots available for priority course', 'error')
            else:
                for _ in range(2):
                    day, start_slot = random.choice(available_slots)
                    classroom = find_available_classroom(day, start_slot)
                    available_slots.remove((day, start_slot))
                    for offset in range(2):
                        new_schedule = Schedule(
                            batch_id=batch.id,
                            course_id=course.id,
                            professor_id=professor.id,
                            day=day,
                            slot=start_slot + offset,
                            classroom_id= classroom.id
                        )
                        db.session.add(new_schedule)
                    try:
                        db.session.commit()
                        flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
                    except:
                        db.session.rollback()
                        flash('Error scheduling priority course (2-hour consecutive)', 'error')
        elif priority and priority_type == "2-1-1":
            if len(available_slots) < credits:
                flash('Not enough available slots for 2-1-1 pattern', 'error')
            else:
                # Step 1: Schedule the 2-hour consecutive slot on one day
                day1, first_slot = available_slots[0]
                classroom = find_available_classroom(day1,first_slot)
                for offset in range(2):  # Schedule 2 consecutive slots
                    new_schedule = Schedule(
                        batch_id=batch.id,
                        course_id=course.id,
                        professor_id=professor.id,
                        day=day1,
                        slot=first_slot + offset,
                        classroom_id= classroom.id
                    )
                    db.session.add(new_schedule)

                # Step 2: Find remaining available slots on different days for 1-hour classes
                day_count = 0
                for day, slot in available_slots[1:]:  # Skip the first day used for 2-hour slot
                    classroom = find_available_classroom(day,slot)
                    if day != day1 and day_count < 2:  # Ensure different days and only 2 additional slots
                        new_schedule = Schedule(
                            batch_id=batch.id,
                            course_id=course.id,
                            professor_id=professor.id,
                            day=day,
                            slot=slot,
                            classroom_id= classroom.id
                        )
                        db.session.add(new_schedule)
                        day_count += 1

                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-1-1)', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling priority course (2-1-1)', 'error')
                else:
                    flash('Required slots are not available for the 2-1-1 pattern', 'error')
        elif priority and priority_type == "2-1":
            if len(available_slots) < credits:
                flash('Not enough available slots for 2-1 pattern', 'error')
            else:
                # Step 1: Schedule the 2-hour consecutive slot on one day
                day1, first_slot = available_slots[0]
                classroom = find_available_classroom(day1, first_slot)
                for offset in range(2):  # Schedule 2 consecutive slots
                    new_schedule = Schedule(
                        batch_id=batch.id,
                        course_id=course.id,
                        professor_id=professor.id,
                        day=day1,
                        slot=first_slot + offset,
                        classroom_id= classroom.id              

                    )
                    db.session.add(new_schedule)

                # Step 2: Find remaining available slots on different days for 1-hour classes
                day_count = 0
                for day, slot in available_slots[1:]:  # Skip the first day used for 2-hour slot
                    classroom = find_available_classroom(day,slot)
                    if day != day1 and day_count < 1:  # Ensure different days and only 1 additional slots
                        new_schedule = Schedule(
                            batch_id=batch.id,
                            course_id=course.id,
                            professor_id=professor.id,
                            day=day,
                            slot=slot,
                            classroom_id= classroom.id
                        )
                        db.session.add(new_schedule)
                        day_count += 1

                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-1)', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling priority course (2-1)', 'error')
                else:
                    flash('Required slots are not available for the 2-1 pattern', 'error')
        elif is_lab:
            # Lab scheduling logic remains the same
            if not available_slots:
                flash('Not enough available slots for lab', 'error')
            else:
                day, start_slot = random.choice(available_slots)
                lab = find_available_lab(day, start_slot) 
                for offset in range(3):
                    new_schedule = Schedule(
                        batch_id=batch.id,
                        course_id=course.id,
                        professor_id=professor.id,
                        day=day,
                        slot=start_slot + offset,
                        lab_id = lab.id

                    )
                    db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Lab scheduled successfully', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling lab', 'error')

        else:
            if len(available_slots) < credits:
                flash('Not enough available slots', 'error')
            else:
                scheduled_slots = random.sample(available_slots, credits)
                for day, slot in scheduled_slots:
                    classroom = find_available_classroom(day, slot)
                    new_schedule = Schedule(
                        batch_id=batch.id,
                        course_id=course.id,
                        professor_id=professor.id,
                        day=day,
                        slot=slot,
                        classroom_id= classroom.id
                    )
                    db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Course scheduled successfully', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling course', 'error')

        return redirect(url_for('manage_batch', batch_id=batch_id))
    schedules = Schedule.query.filter_by(batch_id=batch_id).all()
    timetable = [["-" for _ in range(9)] for _ in range(5)]
    

    for schedule in schedules:
        course = Course.query.get(schedule.course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom= Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        # Construct the timetable entry dynamically
        entry = f"{course.name}({professor.name})"
        if lab is not None:
            entry += f", {lab.name}"
        if classroom is not None:
            entry += f" {classroom.name}"
        # Assign the constructed entry to the timetable
        timetable[schedule.day][schedule.slot] = entry
    return render_template('manage_batch.html', batch=batch, timetable=timetable)
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
